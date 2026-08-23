# Split csv file into multiple parts

import os
from io import BytesIO
import uuid
from typing import Iterator, List
from db.models.knowledgebase.file import KbFileEntity
from db.models.knowledgebase.file_task import KbFileTaskEntity
import pandas as pd
from rag.convert.office_converter import convert_xls_to_xlsx
from rag.split.constants import MAX_PART_ROW_NUM
from pairag.file.store.file_store_helper import file_store
from openpyxl import load_workbook
from loguru import logger

def _create_file_task(
    file_entity: KbFileEntity,
    header: List[str],
    current_rows: List[List[str]],
    current_part: int,
    base_path: str,
) -> KbFileTaskEntity:
    df = pd.DataFrame(current_rows, columns=header)
    file_part_path = f"{base_path}_Part{current_part:04d}.xlsx"
    binary_buffer = BytesIO()
    df.to_excel(binary_buffer, index=False)
    binary_buffer.seek(0)
    upload_result = file_store.write(file=binary_buffer, file_name=file_entity.file_name, file_path=file_part_path, tenant_id=file_entity.tenant_id)
    logger.info(f"Created excel part file: {upload_result.file_path} with {len(current_rows)} rows.")
    return KbFileTaskEntity(
        id=uuid.uuid4().hex,
        file_id=file_entity.id,
        kb_id=file_entity.kb_id,
        file_part=current_part,
        file_path=upload_result.file_path,
        file_version=file_entity.file_version,
        tenant_id=file_entity.tenant_id,
    )

def split_excel(file_entity: KbFileEntity) -> Iterator[KbFileTaskEntity]:
    logger.info(f"Start splitting excel file: {file_entity.file_path}")
    file = file_store.read(file_path=file_entity.file_path, tenant_id=file_entity.tenant_id)
    base_path, ext = os.path.splitext(file_entity.file_path)
    if ext == ".xls":
        file = convert_xls_to_xlsx(file)

    wb = load_workbook(file, data_only=True)

    if len(wb.worksheets) < 1:
        logger.warning("No sheet found in file: {}", file_entity.file_path)
        return

    # 默认读取第一个sheet
    sheet = wb.worksheets[0]
    if sheet.max_row < 1:
        logger.warning("No data found in sheet: {}", sheet.title)
        return

    # 读取表头（第1行）
    header = [str(cell.value) for cell in sheet[1]]
    current_rows = []
    current_part = 1
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        if len(current_rows) > 0 and len(current_rows) % MAX_PART_ROW_NUM == 0:
            yield _create_file_task(file_entity, header, current_rows, current_part, base_path)
            current_rows = []
            current_part += 1
        current_rows.append([str(cell) if cell is not None else "" for cell in row])

    if current_part == 1:
        yield KbFileTaskEntity(
            id=uuid.uuid4().hex,
            file_id=file_entity.id,
            kb_id=file_entity.kb_id,
            file_part=0,
            file_path=file_entity.file_path,
            file_version=file_entity.file_version,
            tenant_id=file_entity.tenant_id,
        )
    else:
        yield _create_file_task(file_entity, header, current_rows, current_part, base_path)

    wb.close()
    logger.info(f"Finished splitting excel file: {file_entity.file_path} into {current_part} parts.")
