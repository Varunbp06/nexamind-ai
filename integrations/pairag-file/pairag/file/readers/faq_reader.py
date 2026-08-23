"""Tabular parser-Excel parser for FAQ file。

"""

import os
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional
from loguru import logger
from openpyxl import load_workbook

import pandas as pd
from llama_index.core.schema import Document
from pairag.file.readers.base import BaseReader
from pairag.file.models.file_item import FileItem


class FAQReader(BaseReader):

    def __init__(
        self,
        *args: Any,
        header_index_max: Optional[int] = 0,
        question_column_index: Optional[int] = 0,
        answer_column_index: Optional[int] = 1,
        **kwargs: Any,
    ) -> None:
        """Init params."""
        super().__init__(*args, **kwargs)
        self._question_column_index = question_column_index if question_column_index is not None else 0
        self._answer_column_index = answer_column_index if answer_column_index is not None else 1
        self._header_index_max = header_index_max  # Allow None to indicate no header row
        # When header_index_max is None, pandas will use numeric column indices (0, 1, 2, ...)
        # Use list of rows from 0 to header_index_max as MultiIndex column names
        if self._header_index_max is None:
            self._pandas_config = {'header': None}
        else:
            self._pandas_config = {'header': list(range(self._header_index_max + 1))}

    def read_xlsx(
        self,
        file: BinaryIO,
        file_extension: Optional[str] = None,
    ):
        """Parse Excel file (supports both .xls and .xlsx with merge_cells handling)."""
        file.seek(0)
        
        if file_extension and file_extension.lower() == ".xls":
            df_temp = pd.read_excel(file, sheet_name=0, engine='xlrd')
            xlsx_file = BytesIO()
            df_temp.to_excel(xlsx_file, engine='openpyxl', index=False)
            xlsx_file.seek(0)
            file = xlsx_file
        
        excel = pd.ExcelFile(load_workbook(file, data_only=True), engine="openpyxl")
        sheet_name = excel.sheet_names[0]
        sheet = excel.book[sheet_name]
        df = excel.parse(sheet_name, **self._pandas_config)

        for item in sheet.merged_cells:
            top_col, top_row, bottom_col, bottom_row = item.bounds
            base_value = item.start_cell.value
            # Convert 1-based index to 0-based index
            top_row -= 1
            top_col -= 1
            # Since the previous lines are set as headers, the coordinates need to be adjusted here.
            if self._header_index_max is not None and self._header_index_max > 0:
                top_row -= self._header_index_max + 1
                bottom_row -= self._header_index_max + 1

            df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value
        return df

    
    
    def _process_dataframe(self, df: pd.DataFrame, extra_info: Optional[Dict] = None, file_name: Optional[str] = None) -> List[Document]:
        """Process DataFrame and create FAQ documents."""

        # Get question and answer columns by index
        if len(df.columns) <= self._question_column_index:
            raise ValueError(f"Question column index {self._question_column_index} is out of range. DataFrame has {len(df.columns)} columns.")
        if len(df.columns) <= self._answer_column_index:
            raise ValueError(f"Answer column index {self._answer_column_index} is out of range. DataFrame has {len(df.columns)} columns.")
        
        question_column = df.columns[self._question_column_index]
        answer_column = df.columns[self._answer_column_index]

        # Build documents for each row
        docs = []
        extra_info = extra_info or {}
        
        for i, row in df.iterrows():
            question = str(row[question_column]) if pd.notna(row[question_column]) else ""
            answer = str(row[answer_column]) if pd.notna(row[answer_column]) else ""
            
            if not question.strip() and not answer.strip():
                continue

            # Build clean embedding text: question only, or question + answer.
            # Avoid "问题:" / "答案:" prefixes — they add noise and hurt
            # semantic matching precision.
            if answer.strip():
                chunk_text = f"{question}\n{answer}"
            else:
                chunk_text = question
            
            row_metadata = extra_info.copy()
            row_metadata["row_number"] = i + 1
            row_metadata["question"] = question
            row_metadata["answer"] = answer
            row_metadata["is_faq"] = True  # signal to skip file_name/title prefix in embedding
            
            docs.append(Document(text=chunk_text, metadata=row_metadata))

        file_display_name = file_name if file_name else "file"
        logger.info(f"Parsed workbook {file_display_name} into {len(docs)} FAQ documents.")
        return docs

    def read(self, file_item: FileItem) -> List[Document]:
        """Read Excel file from FileItem."""
        extra_info = file_item.metadata()
        
        # Use file_item.file directly, unified handling for both .xls and .xlsx
        file_item.file.seek(0)
        df = self.read_xlsx(file_item.file, file_item.file_extension)
        
        return self._process_dataframe(df, extra_info, file_item.file_name)
        