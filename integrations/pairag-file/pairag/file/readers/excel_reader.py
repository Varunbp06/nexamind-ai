"""Tabular parser-Excel parser.

Contains parsers for tabular data files.

"""

import os
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional
from loguru import logger
from openpyxl import load_workbook
from tqdm import tqdm
import pandas as pd
from llama_index.core.schema import Document
from pairag.file.readers.base import BaseReader
from pairag.file.models.file_item import FileItem
from pairag.file.utils.text_utils import replace_consecutive_spaces

class ExcelReader(BaseReader):

    def __init__(
        self,
        *args: Any,
        concat_rows: Optional[bool] = False,
        row_joiner: Optional[str] = "\n",
        header_index_max: Optional[int] = 0,
        format_sheet_data_to_json: Optional[bool] = False,
        sheet_column_filters: Optional[List[str]] = None,
        **kwargs: Any,
    ) -> None:
        """Init params."""
        super().__init__(*args, **kwargs)
        self._concat_rows = concat_rows if concat_rows is not None else False
        self._row_joiner = row_joiner  if row_joiner is not None else "\n"
        self._header_index_max = header_index_max  # Allow None to indicate no header row
        self._format_sheet_data_to_json = format_sheet_data_to_json if format_sheet_data_to_json is not None else False
        self._sheet_column_filters = sheet_column_filters if sheet_column_filters is not None else None
        # Use list of rows from 0 to header_index_max as MultiIndex column names
        if self._header_index_max is None:
            self._pandas_config = {'header': None}
        else:
            self._pandas_config = {'header': list(range(self._header_index_max + 1))}

    def read_xlsx_gen(
        self,
        file: BinaryIO,
        file_extension: Optional[str] = None,
    ):
        """Parse Excel file (supports both .xls and .xlsx with merge_cells handling)."""
        file.seek(0)

        if file_extension and file_extension.lower() == ".xls":
            excel = pd.ExcelFile(file, engine="xlrd")
        else:
            excel = pd.ExcelFile(load_workbook(file, data_only=True), engine="openpyxl")

        sheet_names = list(excel.sheet_names)
        logger.info(f"Parsed Excel file, sheet names: {sheet_names}")
        for sheet_name in tqdm(sheet_names, desc="Parsing excel sheets"):
            sheet = excel.book[sheet_name]
            df = excel.parse(sheet_name, **self._pandas_config)

            for item in sheet.merged_cells:
                top_col, top_row, bottom_col, bottom_row = item.bounds
                base_value = item.start_cell.value
                # Convert 1-based index to 0-based index
                top_row -= 1
                top_col -= 1
                # Since the previous lines are set as headers, the coordinates need to be adjusted here.
                if self._header_index_max is not None and self._header_index_max > 0:
                    top_row -= self._header_index_max + 1
                    bottom_row -= self._header_index_max + 1

                df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value

            df["sheet_name"] = sheet_name
            yield sheet_name, df


    def read(self, file_item: FileItem) -> List[Document]:
        """Read Excel file from FileItem."""
        extra_info = file_item.metadata()
        docs = []
        for sheet_name, df in self.read_xlsx_gen(file_item.file, file_item.file_extension):
            logger.info(f"Parsing excel sheet {sheet_name}")
            
            if self._sheet_column_filters:
                df = df[self._sheet_column_filters]

            # Handle MultiIndex column names by joining them with separator
            def format_column_name(col):
                if isinstance(col, tuple):
                    # MultiIndex column: join with space
                    return " ".join(str(c) for c in col if pd.notna(c) and str(c).strip())
                else:
                    return str(col)

            if self._format_sheet_data_to_json:
                text_list = df.apply(
                    lambda row: str(dict(zip(
                        [format_column_name(col) for col in df.columns], 
                        [str(v) if pd.notna(v) else '' for v in row]
                    ))), axis=1
                ).tolist()
            else:
                text_list = [
                    "\n".join([
                        f"{format_column_name(k)}:{str(v) if pd.notna(v) else ''}" 
                        for k, v in record.items()
                    ])
                    for record in df.to_dict("records")
                ]

            for text in text_list:
                text = replace_consecutive_spaces(text)

            if self._concat_rows:
                logger.info(f"Parsed workbook {file_item.file_name} sheet {sheet_name} into single document.")
                docs.append(
                    Document(
                        id_=file_item.id,
                        text=(self._row_joiner).join(text_list), metadata=extra_info or {}
                    )
                )
            else:
                extra_info = extra_info or {}
                for i, text in enumerate(text_list):
                    row_metadata = extra_info.copy()
                    row_metadata["row_number"] = i + 1
                    docs.append(Document(id_=file_item.id, text=text, metadata=row_metadata))

                logger.info(f"Parsed workbook {file_item.file_name} sheet {sheet_name} into {len(text_list)} documents.")

        return docs